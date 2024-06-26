import pandas as pd
from fpdf import FPDF
import openpyxl
import glob
from pathlib import Path

# This method gets all the archives of the path with the specified extension
filepaths = glob.glob("invoices/*.xlsx")

for filepath in filepaths:
    pdf = FPDF(orientation="P", unit="mm", format="A4")

    pdf.add_page()

    # This function from pathlib extract the filename of the filepath
    filename = Path(filepath).stem

    # Gets the invoice number and the date
    invoice_nr, date = filename.split("-")

    # Way that I knew and used to extract the text from the filename
    # text = f"Invoice nr.{filepath[9:14]}"
    # date_text = f"Date {filepath[15:24]}"

    # Add the text of invoice number and date
    pdf.set_font(family="Times", size=16, style="B")
    pdf.cell(w=0, h=10, txt=f"Invoice nr.{invoice_nr}", ln=1, align="L")
    pdf.cell(w=0, h=10, txt=f"Date {date}", ln=1, align="L")

    df = pd.read_excel(filepath, sheet_name="Sheet 1")
    # Add table header
    columns = list(df.columns)
    columns = [item.replace("_", " ").title() for item in columns]
    pdf.set_font(family="Times", size=12)
    pdf.cell(w=30, h=8, txt=columns[0], border=1)
    pdf.cell(w=60, h=8, txt=columns[1], border=1)
    pdf.cell(w=40, h=8, txt=columns[2], border=1)
    pdf.cell(w=30, h=8, txt=columns[3], border=1)
    pdf.cell(w=30, h=8, txt=columns[4], border=1, ln=1)

    # Add table rows
    for index, row in df.iterrows():
        pdf.set_font(family="Times", size=10)
        pdf.set_text_color(80, 80, 80)
        pdf.cell(w=30, h=8, txt=str(row["product_id"]), border=1)
        pdf.cell(w=60, h=8, txt=str(row["product_name"]), border=1)
        pdf.cell(w=40, h=8, txt=str(row["amount_purchased"]), border=1)
        pdf.cell(w=30, h=8, txt=str(row["price_per_unit"]), border=1)
        pdf.cell(w=30, h=8, txt=str(row["total_price"]), border=1, ln=1)

    # Add total price row
    total_price = df["total_price"].sum()
    pdf.set_font(family="Times", size=10)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(w=30, h=8, border=1)
    pdf.cell(w=60, h=8, border=1)
    pdf.cell(w=40, h=8, border=1)
    pdf.cell(w=30, h=8, border=1)
    pdf.cell(w=30, h=8, txt=str(total_price), border=1, ln=1)

    # Add footer
    pdf.set_font(family="Times", size=14, style="B")
    pdf.set_text_color(0, 0, 0)
    pdf.cell(w=0, h=10, txt=f"The total due amount is {total_price} Euros.",
             ln=1, align="L")
    pdf.cell(w=25, h=10, txt=f"PythonHow", align="L")
    pdf.image("pythonhow.png", w=10)

    pdf.output(f"PDFS/{filename}.pdf")

